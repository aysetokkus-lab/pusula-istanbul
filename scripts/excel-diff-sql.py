"""
Pusula Istanbul - Excel'deki degisiklikleri DB ile karsilastirir, UPDATE/INSERT SQL'i uretir.

Kullanim: python3 excel-diff-sql.py <excel-yolu> <cikti-sql-yolu>

Mantik:
- Excel'de mekan_id var, DB'de var → her alan icin diff, sadece degisenler icin UPDATE
- Excel'de mekan_id var, DB'de yok → INSERT
- Excel'de mekan_id bos ama isim var → INSERT (mekan_id otomatik uretilir)
- Excel'de hucre BOS → "dokunma" (mevcut DB degeri korunur)
- Excel'de hucre "NULL" → SET NULL (alan silinir)
"""
import json
import os
import re
import sys
import urllib.request
import urllib.parse
from openpyxl import load_workbook

# DB'deki kolon adlarinin Excel'deki kolon basliklarina haritalanmasi
KOLON_MAP = {
    'Mekan ID':           'mekan_id',
    'Isim':               'isim',
    'Tip':                'tip',
    'Kategori':           'kategori',
    'Aktif':              'aktif',
    'Mevsimsel':          'mevsimsel',
    'Restorasyon':        'restorasyon',
    'Restorasyon Notu':   'restorasyon_notu',
    'Acilis':             'acilis',
    'Kapanis':            'kapanis',
    'Gise Kapanis':       'gise_kapanis',
    'Yaz Acilis':         'yaz_acilis',
    'Yaz Kapanis':        'yaz_kapanis',
    'Yaz Gise Kapanis':   'yaz_gise_kapanis',
    'Kis Acilis':         'kis_acilis',
    'Kis Kapanis':        'kis_kapanis',
    'Kis Gise Kapanis':   'kis_gise_kapanis',
    'Kapali Gun':         'kapali_gun',
    'Haftasonu Acilis':   'haftasonu_acilis',
    'Haftasonu Kapanis':  'haftasonu_kapanis',
    'Cuma Kapali Bas':    'cuma_kapali_bas',
    'Cuma Kapali Bit':    'cuma_kapali_bit',
    'Fiyat Yerli':        'fiyat_yerli',
    'Fiyat Yabanci':      'fiyat_yabanci',
    'Fiyat Indirimli':    'fiyat_indirimli',
    'MuzeKart':           'muzekart',
    'Ozel Not':           'ozel_not',
    'Resmi Site':         'site',
    'Kaynak':             'kaynak',
}

# Bool tipindeki alanlar
BOOL_ALANLAR = {'aktif', 'mevsimsel', 'restorasyon'}

# Integer tipindeki alanlar
INT_ALANLAR = {'kapali_gun'}

# Saat tipindeki alanlar (Postgres'te time olarak gidecek ama string formatinda da kabul edilir)
SAAT_ALANLAR = {
    'acilis', 'kapanis', 'gise_kapanis',
    'yaz_acilis', 'yaz_kapanis', 'yaz_gise_kapanis',
    'kis_acilis', 'kis_kapanis', 'kis_gise_kapanis',
    'haftasonu_acilis', 'haftasonu_kapanis',
    'cuma_kapali_bas', 'cuma_kapali_bit',
}

def env_oku(yol):
    d = {}
    if not os.path.exists(yol): return d
    with open(yol, 'r', encoding='utf-8') as f:
        for satir in f:
            satir = satir.strip()
            if not satir or satir.startswith('#') or '=' not in satir: continue
            k, v = satir.split('=', 1)
            d[k.strip()] = v.strip().strip('"').strip("'")
    return d

def db_kayitlari_cek(url, key):
    rest_url = f"{url}/rest/v1/mekan_saatleri?" + urllib.parse.urlencode({'select': '*'})
    req = urllib.request.Request(rest_url, headers={
        'apikey': key, 'Authorization': f'Bearer {key}', 'Accept': 'application/json',
    })
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode('utf-8'))

def excel_oku(yol):
    """Excel'deki Veri sayfasini oku, kayit listesi dondur."""
    wb = load_workbook(yol, data_only=True)
    ws = wb['Veri']

    # Basliklar
    basliklar = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # Excel basliklarini DB anahtarlarina cevir
    db_anahtarlari = [KOLON_MAP.get(b) for b in basliklar]

    kayitlar = []
    for satir_idx in range(2, ws.max_row + 1):
        kayit = {}
        bos_satir = True
        for col_idx, db_anahtar in enumerate(db_anahtarlari, start=1):
            if db_anahtar is None: continue
            val = ws.cell(satir_idx, col_idx).value
            kayit[db_anahtar] = val
            if val is not None and (not isinstance(val, str) or val.strip() != ''):
                bos_satir = False
        if not bos_satir:
            kayit['_satir_no'] = satir_idx
            kayitlar.append(kayit)
    return kayitlar

def excel_degeri_normalize(deger, alan):
    """Excel'den gelen degeri Python'a uygun hale getir."""
    if deger is None: return None
    if isinstance(deger, str):
        s = deger.strip()
        if s == '': return None
        if s.upper() == 'NULL': return 'EXPLICIT_NULL'  # Ozel sentinel — alan silinmek isteniyor
        if alan in BOOL_ALANLAR:
            if s.upper() in ('TRUE', '1', 'YES', 'EVET', 'TRUE'): return True
            if s.upper() in ('FALSE', '0', 'NO', 'HAYIR', 'FALSE'): return False
        if alan in INT_ALANLAR:
            try: return int(s)
            except: return s
        return s
    if isinstance(deger, bool): return deger
    if isinstance(deger, int) and alan in INT_ALANLAR: return deger
    if isinstance(deger, float):
        if alan in INT_ALANLAR: return int(deger)
        if deger.is_integer(): return str(int(deger))
        return str(deger)
    return deger

def db_degeri_normalize(deger, alan):
    """DB'den gelen degeri karsilastirma icin uygun hale getir."""
    if deger is None: return None
    # Saatler PostgreSQL time tipinde "HH:MM:SS" olarak gelebilir, "HH:MM" olabilir
    if alan in SAAT_ALANLAR and isinstance(deger, str):
        # "09:00:00" → "09:00"
        m = re.match(r'^(\d{2}:\d{2})(:\d{2})?$', deger)
        if m: return m.group(1)
    return deger

def deger_eslesir_mi(excel_d, db_d, alan):
    """Iki deger esit mi? (None == None ya da string esit)"""
    if excel_d is None and db_d is None: return True
    if excel_d is None or db_d is None: return False
    if excel_d == 'EXPLICIT_NULL': return db_d is None  # NULL set istegi: db zaten null mi
    # String karsilastirma
    return str(excel_d).strip() == str(db_d).strip()

def sql_value(deger, alan):
    """Python degerini SQL literal'ina cevir."""
    if deger is None or deger == 'EXPLICIT_NULL': return 'NULL'
    if alan in BOOL_ALANLAR:
        if isinstance(deger, bool): return 'TRUE' if deger else 'FALSE'
        if isinstance(deger, str):
            if deger.upper() == 'TRUE': return 'TRUE'
            if deger.upper() == 'FALSE': return 'FALSE'
    if alan in INT_ALANLAR:
        return str(int(deger))
    s = str(deger).replace("'", "''")
    return f"'{s}'"

def update_sql_uret(mekan_id, degisikler):
    """UPDATE statement uret."""
    if not degisikler: return None
    set_parcalari = []
    for alan, yeni_deger in degisikler.items():
        set_parcalari.append(f"  {alan} = {sql_value(yeni_deger, alan)}")
    set_parcalari.append(f"  guncelleme_tarihi = NOW()")
    safe_id = mekan_id.replace("'", "''")
    return f"UPDATE mekan_saatleri SET\n" + ",\n".join(set_parcalari) + f"\nWHERE mekan_id = '{safe_id}';"

def insert_sql_uret(kayit):
    """INSERT statement uret."""
    # Yeni mekan icin mekan_id otomatik uretebilir (isimden)
    if not kayit.get('mekan_id'):
        if not kayit.get('isim'):
            return None, "isim bos, mekan_id uretilemez"
        # Basit slug
        slug = kayit['isim'].lower()
        slug = re.sub(r'[^a-z0-9ğüşıöç]+', '_', slug)
        slug = slug.strip('_')
        kayit['mekan_id'] = slug

    # aktif_mevsim default 'kis'
    kayit.setdefault('aktif_mevsim', 'kis')

    alanlar = []
    degerler = []
    for alan, deger in kayit.items():
        if alan.startswith('_'): continue
        if deger is None or deger == '': continue
        alanlar.append(alan)
        degerler.append(sql_value(deger, alan))

    return f"INSERT INTO mekan_saatleri ({', '.join(alanlar)})\nVALUES ({', '.join(degerler)});", None

def main():
    if len(sys.argv) < 3:
        print('Kullanim: python3 excel-diff-sql.py <excel-yolu> <cikti-sql-yolu>')
        sys.exit(1)

    excel_yolu = sys.argv[1]
    cikti_yolu = sys.argv[2]

    # .env oku
    proje_kok = '/sessions/magical-upbeat-mccarthy/mnt/istanbul-rehber'
    env = env_oku(os.path.join(proje_kok, '.env'))
    url = env.get('SUPABASE_URL') or env.get('EXPO_PUBLIC_SUPABASE_URL') or 'https://rzlfghjpsximthlolfxo.supabase.co'
    key = env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not key:
        print("HATA: SUPABASE_SERVICE_ROLE_KEY bulunamadi"); sys.exit(1)

    print(f"Excel okunuyor: {excel_yolu}")
    excel_kayitlari = excel_oku(excel_yolu)
    print(f"  {len(excel_kayitlari)} satir okundu")

    print(f"DB'den kayitlar cekiliyor...")
    db_kayitlari_list = db_kayitlari_cek(url, key)
    db_index = {k['mekan_id']: k for k in db_kayitlari_list}
    print(f"  {len(db_kayitlari_list)} kayit DB'de var")

    # Diff hesapla
    update_sqlleri = []
    insert_sqlleri = []
    update_ozet = []
    insert_ozet = []
    atlanan = []

    for excel_kayit in excel_kayitlari:
        mekan_id = excel_kayit.get('mekan_id')
        satir_no = excel_kayit['_satir_no']

        if mekan_id and mekan_id in db_index:
            # UPDATE potansiyeli
            db_kayit = db_index[mekan_id]
            degisikler = {}
            for alan in KOLON_MAP.values():
                excel_deger_raw = excel_kayit.get(alan)
                excel_deger = excel_degeri_normalize(excel_deger_raw, alan)
                db_deger = db_degeri_normalize(db_kayit.get(alan), alan)

                # Bos hucre = "dokunma"
                if excel_deger is None and excel_deger_raw is None:
                    continue
                if excel_deger == 'EXPLICIT_NULL':
                    if db_deger is not None:
                        degisikler[alan] = None
                    continue
                if not deger_eslesir_mi(excel_deger, db_deger, alan):
                    degisikler[alan] = excel_deger

            if degisikler:
                sql = update_sql_uret(mekan_id, degisikler)
                update_sqlleri.append(sql)
                update_ozet.append((mekan_id, db_kayit.get('isim'), list(degisikler.keys())))
            else:
                atlanan.append((mekan_id, db_kayit.get('isim'), 'Degisiklik yok'))
        else:
            # INSERT
            kayit_kopya = {k: v for k, v in excel_kayit.items() if not k.startswith('_')}
            # None'lari ve bos string'leri temizle
            temiz = {}
            for alan, deger in kayit_kopya.items():
                normalize = excel_degeri_normalize(deger, alan)
                if normalize is not None and normalize != 'EXPLICIT_NULL':
                    temiz[alan] = normalize
            sql, hata = insert_sql_uret(temiz)
            if sql:
                insert_sqlleri.append(sql)
                insert_ozet.append((temiz.get('mekan_id'), temiz.get('isim')))
            else:
                atlanan.append((mekan_id or '?', temiz.get('isim') or '?', f"Atlandi: {hata}"))

    # SQL dosyasini yaz
    with open(cikti_yolu, 'w', encoding='utf-8') as f:
        f.write("-- Pusula Istanbul - Mekan Saatleri Toplu Guncelleme\n")
        f.write(f"-- Kaynak: {os.path.basename(excel_yolu)}\n")
        f.write(f"-- Olusturulma: $(date)\n")
        f.write(f"-- {len(update_sqlleri)} UPDATE + {len(insert_sqlleri)} INSERT\n\n")
        f.write("BEGIN;\n\n")
        if update_sqlleri:
            f.write("-- ═══════════════ UPDATE'ler ═══════════════\n\n")
            for sql in update_sqlleri:
                f.write(sql + "\n\n")
        if insert_sqlleri:
            f.write("-- ═══════════════ INSERT'ler ═══════════════\n\n")
            for sql in insert_sqlleri:
                f.write(sql + "\n\n")
        f.write("COMMIT;\n")

    # Rapor
    print()
    print("═" * 70)
    print(f"OZET")
    print("═" * 70)
    print(f"  UPDATE: {len(update_sqlleri)} kayit")
    print(f"  INSERT: {len(insert_sqlleri)} kayit")
    print(f"  Atlanan/degisiklik yok: {len(atlanan)} kayit")
    print()

    if update_ozet:
        print("DEGISEN KAYITLAR (UPDATE):")
        for mid, isim, alanlar in update_ozet:
            print(f"  • {isim} ({mid})")
            for a in alanlar:
                print(f"      - {a}")
        print()

    if insert_ozet:
        print("YENI KAYITLAR (INSERT):")
        for mid, isim in insert_ozet:
            print(f"  • {isim} (id: {mid})")
        print()

    print(f"SQL dosyasi: {cikti_yolu}")
    print(f"Toplam SQL satiri: {len(update_sqlleri) + len(insert_sqlleri)}")

if __name__ == '__main__':
    main()
