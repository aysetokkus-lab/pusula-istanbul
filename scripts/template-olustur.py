"""
Pusula Istanbul - Mekan Saatleri Veri Giris Template Olusturucu
Kullanim: python3 template-olustur.py [cikti_yolu]
Default: /Users/aysetokkus/istanbul-rehber/mekan-saatleri-veri-giris.xlsx
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Pusula renk paleti
RENK = {
    'koyu_mavi':   '005A8D',  # zorunlu kolonlar
    'orta_mavi':   '0077B6',  # ana baslik
    'acik_mavi':   '48CAE4',  # saat kolonlari
    'cok_acik':    'CAF0F8',  # mevsimsel saat
    'altin':       'C77A15',  # fiyat
    'altin_acik':  'F4D58D',  # fiyat ack
    'mor':         '7B2D8E',  # meta
    'mor_acik':    'E6D7EE',  # meta acik
    'amber':       'E09F3E',  # durum
    'gri':         'E5E7EB',  # ornek satir bg
    'beyaz':       'FFFFFF',
}

KOLONLAR = [
    # (header, key, group, width, comment, validation)
    ('Mekan ID',       'mekan_id',          'kimlik', 18,
     "Benzersiz ID. Yeni mekan icin bos birakabilirsin (otomatik isimden uretilir). Mevcut mekani guncellemek icin tam ID'yi yaz (orn: 'topkapi', 'galata_kulesi').",
     None),
    ('Isim',           'isim',              'kimlik', 32,
     "Mekanin gosterim adi. Turkce karakterli yazabilirsin: 'Topkapi Sarayi', 'Galata Kulesi'.",
     None),
    ('Tip',            'tip',               'kimlik', 12,
     "Mekan tipi. Drop-down'dan sec.",
     '"cami,muze,saray,kasir,ozel_muze,kule"'),
    ('Kategori',       'kategori',          'kimlik', 16,
     "Uygulamada hangi sekmede gorunecek. Drop-down'dan sec. milli_saraylar = 'Milli Saraylar', muzeler = 'Muzeler', ozel_muzeler = 'Ozel Muzeler', camiler = 'Camiler'.",
     '"milli_saraylar,muzeler,ozel_muzeler,camiler"'),

    ('Aktif',          'aktif',             'durum',  10,
     "TRUE = uygulamada gorunsun. FALSE = gizle (gecici kapanis vs.). Drop-down'dan sec.",
     '"TRUE,FALSE"'),
    ('Mevsimsel',      'mevsimsel',         'durum',  12,
     "TRUE = yaz/kis ayri saatler. FALSE = tek saat (sabit acilis/kapanis kullanilir). Drop-down'dan sec.",
     '"TRUE,FALSE"'),
    ('Restorasyon',    'restorasyon',       'durum',  13,
     "TRUE = mekan restorasyonda kapali. Uygulama 'Restorasyon nedeniyle kapali' gosterir. Drop-down'dan sec.",
     '"TRUE,FALSE"'),
    ('Restorasyon Notu', 'restorasyon_notu', 'durum',  30,
     "Restorasyon detayi. Orn: '2026 Mart-Haziran arasi kapali'. Sadece restorasyon=TRUE ise gosterilir.",
     None),

    ('Acilis',         'acilis',            'sabit_saat', 9,
     "Sabit acilis saati (mevsimsel=FALSE icin). Format: 'HH:MM' orn: '09:00'.",
     None),
    ('Kapanis',        'kapanis',           'sabit_saat', 9,
     "Sabit ziyaret bitis saati. Format: 'HH:MM' orn: '17:30'.",
     None),
    ('Gise Kapanis',   'gise_kapanis',      'sabit_saat', 11,
     "Gisenin son bilet sattigi saat. Genelde kapanis -30 dk. Format: 'HH:MM'.",
     None),

    ('Yaz Acilis',     'yaz_acilis',        'mevsim_saat', 10,
     "Yaz mevsimi acilis (1 May - 31 Eki). Sadece mevsimsel=TRUE ise.",
     None),
    ('Yaz Kapanis',    'yaz_kapanis',       'mevsim_saat', 10,
     "Yaz mevsimi ziyaret bitis. Sadece mevsimsel=TRUE ise.",
     None),
    ('Yaz Gise Kapanis', 'yaz_gise_kapanis', 'mevsim_saat', 13,
     "Yaz mevsimi gise kapanis.",
     None),
    ('Kis Acilis',     'kis_acilis',        'mevsim_saat', 10,
     "Kis mevsimi acilis (1 Kas - 30 Nis). Sadece mevsimsel=TRUE ise.",
     None),
    ('Kis Kapanis',    'kis_kapanis',       'mevsim_saat', 10,
     "Kis mevsimi ziyaret bitis. Sadece mevsimsel=TRUE ise.",
     None),
    ('Kis Gise Kapanis', 'kis_gise_kapanis', 'mevsim_saat', 13,
     "Kis mevsimi gise kapanis.",
     None),

    ('Kapali Gun',     'kapali_gun',        'gun_saat', 12,
     "Haftada hangi gun kapali. Drop-down'dan sec. Bos = her gun acik.",
     '"0,1,2,3,4,5,6"'),
    ('Haftasonu Acilis', 'haftasonu_acilis', 'gun_saat', 14,
     "Hafta sonu (Ctesi+Pazar) farkli acilis. Bos = hafta ici ile ayni.",
     None),
    ('Haftasonu Kapanis', 'haftasonu_kapanis', 'gun_saat', 14,
     "Hafta sonu farkli kapanis.",
     None),
    ('Cuma Kapali Bas', 'cuma_kapali_bas',  'gun_saat', 12,
     "Cuma namazi icin kapanis baslangici (camiler icin). Format: 'HH:MM'.",
     None),
    ('Cuma Kapali Bit', 'cuma_kapali_bit',  'gun_saat', 12,
     "Cuma namazi sonrasi tekrar acilis saati.",
     None),

    ('Fiyat Yerli',    'fiyat_yerli',       'fiyat',  13,
     "Yerli ziyaretci ucreti. Orn: '120 TL', 'Ucretsiz'.",
     None),
    ('Fiyat Yabanci',  'fiyat_yabanci',     'fiyat',  13,
     "Yabanci ziyaretci ucreti. Orn: '300 TL', '15 EUR'.",
     None),
    ('Fiyat Indirimli', 'fiyat_indirimli',  'fiyat',  13,
     "Indirimli ucret (ogrenci, 65+). Orn: '60 TL'.",
     None),
    ('MuzeKart',       'muzekart',          'fiyat',  11,
     "MuzeKart gecerli mi? Drop-down'dan sec. 'gecerli' = gecer, 'gecmez' = gecmez, bos = bilgi yok.",
     '"gecerli,gecmez"'),

    ('Ozel Not',       'ozel_not',          'meta',   30,
     "Onemli notlar (orn: 'Bayramda kapali', 'Combined ticket var'). Uygulamada 'Bilgi' kutusunda gosterilir.",
     None),
    ('Resmi Site',     'site',              'meta',   35,
     "Mekanin resmi web sitesi. Tam URL. Orn: 'https://www.millisaraylar.gov.tr/Lokasyon/1/Topkapi-Sarayi'.",
     None),
    ('Kaynak',         'kaynak',            'meta',   18,
     "Verinin kaynagi (orn: 'muze.gov.tr', 'millisaraylar.gov.tr', 'manuel'). Audit icin.",
     None),
]

GRUP_RENGI = {
    'kimlik':      RENK['koyu_mavi'],
    'durum':       RENK['orta_mavi'],
    'sabit_saat':  RENK['acik_mavi'],
    'mevsim_saat': RENK['cok_acik'],
    'gun_saat':    RENK['amber'],
    'fiyat':       RENK['altin'],
    'meta':        RENK['mor'],
}

# Acik renk gruplari icin koyu yazi rengi gerekli
KOYU_YAZI_GRUPLARI = {'mevsim_saat', 'gun_saat'}

def kenarlik(stil='thin'):
    s = Side(border_style=stil, color='9CA3AF')
    return Border(left=s, right=s, top=s, bottom=s)

def main():
    wb = Workbook()

    # ═══════════════ Sayfa 1: Veri ═══════════════
    ws = wb.active
    ws.title = 'Veri'

    # Donmus baslik
    ws.freeze_panes = 'E2'  # ilk satir + ilk 4 kolon (kimlik) donsun

    # Baslik satiri
    for idx, (header, key, group, width, comment, validation) in enumerate(KOLONLAR, start=1):
        col = get_column_letter(idx)
        cell = ws.cell(row=1, column=idx, value=header)
        bg_renk = GRUP_RENGI[group]
        yazi_rengi = '000000' if group in KOYU_YAZI_GRUPLARI else 'FFFFFF'
        cell.fill = PatternFill('solid', start_color=bg_renk)
        cell.font = Font(name='Arial', bold=True, color=yazi_rengi, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = kenarlik('medium')
        ws.column_dimensions[col].width = width

        if comment:
            cell.comment = Comment(comment, 'Pusula')

        if validation:
            dv = DataValidation(type='list', formula1=validation, allow_blank=True, showErrorMessage=True)
            dv.error = f"Gecerli degerler: {validation.strip(chr(34))}"
            dv.errorTitle = 'Gecersiz deger'
            dv.add(f'{col}2:{col}1000')
            ws.add_data_validation(dv)

    ws.row_dimensions[1].height = 38

    # Bos veri satirlari (2. satirdan itibaren 250 satir — ornek satirlar yok, mevcut kayitlar template-doldur.py ile eklenir)
    for satir_idx in range(2, 252):
        for col_idx in range(1, len(KOLONLAR) + 1):
            cell = ws.cell(row=satir_idx, column=col_idx)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = kenarlik('thin')

    # ═══════════════ Sayfa 2: Yardim ═══════════════
    ws2 = wb.create_sheet('Yardim')

    yardim_satirlari = [
        ('PUSULA ISTANBUL — VERI GIRIS REHBERI', 'baslik'),
        ('', None),
        ('Bu dosyayi nasil kullanacaksin', 'altbaslik'),
        ('1. "Veri" sayfasinda mevcut mekanlari guncellemek icin Mekan ID kolonuna tam ID yaz (orn: "topkapi"). Yeni mekan eklemek icin Mekan ID bos birakilabilir.', None),
        ('2. Drop-down\'lu kolonlarda hucreyi tikla, ok isaretine bas, listeden sec.', None),
        ('3. Saat formati her zaman HH:MM olmali (24 saat). Orn: 09:00, 17:30, 23:00.', None),
        ('4. TRUE/FALSE kolonlari tek harf yerine tam yazilmali. Drop-down kullan.', None),
        ('5. Doldurmadiginca alan bos kalsin. Bosluk = NULL = "veri yok" anlamina gelir.', None),
        ('6. Doldurduktan sonra dosyayi kaydet, Pusula\'ya ver. SQL uretilir, Supabase\'de calistirilir.', None),
        ('', None),
        ('Kolon gruplari ve renkleri', 'altbaslik'),
        ('Koyu Mavi (Kimlik): Mekan ID, Isim, Tip, Kategori — DOLDURULMASI ZORUNLU', None),
        ('Mavi (Durum): Aktif, Mevsimsel, Restorasyon, Restorasyon Notu', None),
        ('Acik Mavi (Sabit Saat): Acilis, Kapanis, Gise Kapanis — sadece Mevsimsel=FALSE ise', None),
        ('Cok Acik Mavi (Mevsimsel Saat): Yaz/Kis Acilis/Kapanis/Gise — sadece Mevsimsel=TRUE ise', None),
        ('Amber (Gun-Saat): Kapali Gun, Haftasonu, Cuma kapali saatleri', None),
        ('Altin (Fiyat): Yerli, Yabanci, Indirimli, MuzeKart', None),
        ('Mor (Meta): Ozel Not, Resmi Site, Kaynak', None),
        ('', None),
        ('Tip kolonu gecerli degerleri', 'altbaslik'),
        ('cami    — Camiler (Sultanahmet, Suleymaniye vb.)', None),
        ('muze    — Genel muzeler (Ayasofya, Arkeoloji, TIEM vb.)', None),
        ('saray   — Milli Saray (Topkapi, Dolmabahce, Yildiz vb.)', None),
        ('kasir   — Kucuk saray/kasir (Kucuksu, Aynalikavak, Ihlamur)', None),
        ('ozel_muze — Ozel muzeler (Pera, Sakip Sabanci vb.)', None),
        ('kule    — Kuleler (Galata Kulesi, Kiz Kulesi)', None),
        ('', None),
        ('Kategori kolonu gecerli degerleri', 'altbaslik'),
        ('milli_saraylar — Uygulamada "Milli Saraylar" sekmesi', None),
        ('muzeler        — "Muzeler" sekmesi', None),
        ('ozel_muzeler   — "Ozel Muzeler" sekmesi', None),
        ('camiler        — "Camiler" sekmesi', None),
        ('NOT: Tip ve Kategori farkli olabilir. Orn: Galata Kulesi tip="kule" ama kategori="muzeler".', None),
        ('', None),
        ('Kapali Gun degerleri (haftanin gunleri)', 'altbaslik'),
        ('0 = Pazar', None),
        ('1 = Pazartesi', None),
        ('2 = Sali', None),
        ('3 = Carsamba', None),
        ('4 = Persembe', None),
        ('5 = Cuma', None),
        ('6 = Cumartesi', None),
        ('Bos = Hicbir gun kapali degil', None),
        ('', None),
        ('MuzeKart kolonu gecerli degerleri', 'altbaslik'),
        ('gecerli — MuzeKart ile giris yapilabilir', None),
        ('gecmez  — MuzeKart kabul edilmez (orn: ozel muzeler, Galata Mevlevihanesi)', None),
        ('Bos     — Bilgi yok / uygulanabilir degil (camiler ucretsiz oldugu icin gerek yok)', None),
        ('', None),
        ('Mevsimsel mantigi', 'altbaslik'),
        ('Eger mekan yaz ve kis aylarinda farkli saatlerde aciksa: Mevsimsel=TRUE.', None),
        ('Bu durumda Yaz Acilis/Kapanis/Gise + Kis Acilis/Kapanis/Gise alanlarini doldur.', None),
        ('Sabit Acilis/Kapanis/Gise alanlarini bos birak.', None),
        ('Eger mekan yil boyu ayni saatlerdeyse: Mevsimsel=FALSE.', None),
        ('Bu durumda sadece Sabit Acilis/Kapanis/Gise alanlarini doldur.', None),
        ('Yaz/Kis alanlarini bos birak.', None),
        ('', None),
        ('Mevsim degisikligi (1 Mayis ve 1 Kasim)', 'altbaslik'),
        ('Aktif mevsim DB\'de "aktif_mevsim" alaninda tutulur ve admin panelinden tek tikla degistirilir.', None),
        ('Bu Excel\'den degistirmen GEREKMEZ — Yaz/Kis saatlerini doldur, mevsim gecisini admin panelden yap.', None),
        ('', None),
        ('Cuma kapali saatleri (sadece camiler)', 'altbaslik'),
        ('Cuma namazi icin gecici kapanis saatleri.', None),
        ('Cuma Kapali Bas: kapanis baslangici (orn: 12:30)', None),
        ('Cuma Kapali Bit: tekrar acilis (orn: 14:30)', None),
        ('Sultanahmet Camii\'nin kendine ozgu pencere yapisi var (admin panelinden ayrica yonetilir).', None),
        ('', None),
        ('Cevirisi belirsiz olan alanlar', 'altbaslik'),
        ('Restorasyon Notu: TRUE oldugunda kullaniciya gosterilen aciklama. Bos olabilir.', None),
        ('Ozel Not: Onemli ek bilgi. Combined ticket bilgisi, bayram kapanisi gibi.', None),
        ('Resmi Site: tam URL. Bos olabilir ama tercih edilir (kullanici "Resmi Site" butonu gorur).', None),
        ('Kaynak: Audit icin. "manuel", "muze.gov.tr", "millisaraylar.gov.tr" gibi.', None),
        ('', None),
        ('Yeni mekan eklemek', 'altbaslik'),
        ('Mekan ID bos birakilabilir — isimden otomatik uretilir (orn: "Yeni Saray" → "yeni_saray").', None),
        ('Veya kendin yazabilirsin (kucuk harf, alt cizgi, Turkce karakter olabilir).', None),
        ('Tip ve Kategori\'yi mutlaka doldur — yoksa SQL hata verir.', None),
        ('', None),
        ('Mevcut mekani guncellemek', 'altbaslik'),
        ('Mekan ID kolonunda tam ID yaz (orn: "topkapi", "sultanahmet_camii", "ms_resim").', None),
        ('SQL UPDATE kullanir — sadece doldurdugun alanlar guncellenir, bos birakilan alanlar DEGISMEZ.', None),
        ('Eger bir alani SILMEK istiyorsan ve eskiden dolu ise, "NULL" yazabilirsin (sirf bos degil) — ama bunu kullanmadan once Pusula\'ya sor.', None),
        ('', None),
        ('Mevcut mekan ID listesini almak', 'altbaslik'),
        ('Supabase SQL Editor\'de su sorguyu calistir:', None),
        ('  SELECT mekan_id, isim, tip, kategori, aktif FROM mekan_saatleri ORDER BY kategori, isim;', None),
        ('Cikan ID\'leri Excel\'in Mekan ID kolonuna kopyala, sonra her satirin guncellenmesini istedigin alanlari doldur.', None),
    ]

    for satir_idx, (metin, stil) in enumerate(yardim_satirlari, start=1):
        cell = ws2.cell(row=satir_idx, column=1, value=metin)
        if stil == 'baslik':
            cell.font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color=RENK['orta_mavi'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[satir_idx].height = 30
        elif stil == 'altbaslik':
            cell.font = Font(name='Arial', bold=True, size=12, color=RENK['orta_mavi'])
            cell.fill = PatternFill('solid', start_color='F0F9FF')
            ws2.row_dimensions[satir_idx].height = 22
        else:
            cell.font = Font(name='Arial', size=11, color='374151')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws2.column_dimensions['A'].width = 120
    ws2.row_dimensions[1].height = 35
    ws2.merge_cells('A1:D1')

    # ═══════════════ Sayfa 3: Tatil Gunleri (gelecek) ═══════════════
    ws3 = wb.create_sheet('Tatil Gunleri')

    tatil_kolonlari = [
        ('Mekan ID', 24, "Hangi mekana ozel tatil gunu (orn: 'yildiz_sarayi'). Bos = tum mekanlar."),
        ('Tarih',    14, "Tatil gunu tarihi. Format: YYYY-MM-DD (orn: 2026-04-23)."),
        ('Sebep',    18, "Tatil sebebi. Drop-down: resmi_tatil, dini_tatil, ozel."),
        ('Aciklama', 60, "Aciklama (orn: '23 Nisan Ulusal Egemenlik ve Cocuk Bayrami')."),
    ]

    bilgi_satiri = ws3.cell(row=1, column=1, value=(
        'NOT: Bu sayfa GELECEK ICIN hazirlandi. Su an DB\'de "mekan_tatil_gunleri" tablosu yok. '
        'v1.1.0\'da eklenmesi planlanan ozellik. Doldurabilirsin — schema eklendiginde toplu insert yapilir. '
        'Doldurmasi zorunlu degil, bos birakilabilir.'
    ))
    bilgi_satiri.font = Font(name='Arial', italic=True, size=10, color='B45309')
    bilgi_satiri.fill = PatternFill('solid', start_color='FEF3C7')
    bilgi_satiri.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws3.row_dimensions[1].height = 50
    ws3.merge_cells('A1:D1')

    for idx, (header, width, comment) in enumerate(tatil_kolonlari, start=1):
        col = get_column_letter(idx)
        cell = ws3.cell(row=2, column=idx, value=header)
        cell.fill = PatternFill('solid', start_color=RENK['amber'])
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = kenarlik('medium')
        cell.comment = Comment(comment, 'Pusula')
        ws3.column_dimensions[col].width = width

    ws3.row_dimensions[2].height = 28
    ws3.freeze_panes = 'A3'

    # Sebep dropdown
    dv_sebep = DataValidation(type='list', formula1='"resmi_tatil,dini_tatil,ozel"', allow_blank=True)
    dv_sebep.add('C3:C100')
    ws3.add_data_validation(dv_sebep)

    # Tatil sayfasi 50 bos satir
    for satir_idx in range(3, 53):
        for col_idx in range(1, 5):
            cell = ws3.cell(row=satir_idx, column=col_idx)
            cell.font = Font(name='Arial', size=10)
            cell.border = kenarlik('thin')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Kaydet
    cikti = sys.argv[1] if len(sys.argv) > 1 else '/Users/aysetokkus/istanbul-rehber/mekan-saatleri-veri-giris.xlsx'
    wb.save(cikti)
    print(f"OK: {cikti}")
    print(f"Sayfa sayisi: {len(wb.sheetnames)}")
    print(f"Sayfalar: {', '.join(wb.sheetnames)}")
    print(f"Veri sayfasi kolon sayisi: {len(KOLONLAR)}")

if __name__ == '__main__':
    main()
