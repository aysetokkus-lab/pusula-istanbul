"""
Pusula Istanbul - Template'e mevcut Supabase kayitlarini doldurur.
Kullanim: python3 template-doldur.py <template-yolu>
"""
import json
import os
import sys
import urllib.request
import urllib.parse
from openpyxl import load_workbook

# Excel'deki kolon sirasi (template-olustur.py ile birebir uyumlu)
KOLON_SIRASI = [
    'mekan_id', 'isim', 'tip', 'kategori',
    'aktif', 'mevsimsel', 'restorasyon', 'restorasyon_notu',
    'acilis', 'kapanis', 'gise_kapanis',
    'yaz_acilis', 'yaz_kapanis', 'yaz_gise_kapanis',
    'kis_acilis', 'kis_kapanis', 'kis_gise_kapanis',
    'kapali_gun', 'haftasonu_acilis', 'haftasonu_kapanis',
    'cuma_kapali_bas', 'cuma_kapali_bit',
    'fiyat_yerli', 'fiyat_yabanci', 'fiyat_indirimli', 'muzekart',
    'ozel_not', 'site', 'kaynak',
]

def env_oku(env_yolu):
    """Basit .env parser — anahtar=deger formati."""
    degerler = {}
    if not os.path.exists(env_yolu):
        return degerler
    with open(env_yolu, 'r', encoding='utf-8') as f:
        for satir in f:
            satir = satir.strip()
            if not satir or satir.startswith('#'):
                continue
            if '=' not in satir:
                continue
            anahtar, deger = satir.split('=', 1)
            anahtar = anahtar.strip()
            deger = deger.strip().strip('"').strip("'")
            degerler[anahtar] = deger
    return degerler

def supabase_kayitlari_cek(url, service_key):
    """Supabase REST API ile mekan_saatleri tablosunu cek."""
    params = {
        'select': '*',
        'aktif': 'eq.true',
        'order': 'kategori,isim',
    }
    rest_url = f"{url}/rest/v1/mekan_saatleri?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(rest_url, headers={
        'apikey': service_key,
        'Authorization': f'Bearer {service_key}',
        'Accept': 'application/json',
    })
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode('utf-8'))

def deger_donustur(deger):
    """Python tipini Excel'e uygun string'e cevir."""
    if deger is None:
        return ''
    if isinstance(deger, bool):
        return 'TRUE' if deger else 'FALSE'
    if isinstance(deger, int):
        return str(deger)
    return deger

def main():
    if len(sys.argv) < 2:
        print('Kullanim: python3 template-doldur.py <template-yolu>')
        sys.exit(1)

    template_yolu = sys.argv[1]
    proje_kok = os.path.dirname(os.path.abspath(template_yolu))
    env_yolu = os.path.join(proje_kok, '.env')

    env = env_oku(env_yolu)
    url = env.get('SUPABASE_URL') or env.get('EXPO_PUBLIC_SUPABASE_URL') or 'https://rzlfghjpsximthlolfxo.supabase.co'
    key = env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not key:
        print(f"HATA: SUPABASE_SERVICE_ROLE_KEY .env'de bulunamadi ({env_yolu})")
        sys.exit(1)

    print(f"Supabase'den kayitlar cekiliyor: {url}")
    kayitlar = supabase_kayitlari_cek(url, key)
    print(f"  {len(kayitlar)} kayit alindi")

    print(f"Template aciliyor: {template_yolu}")
    wb = load_workbook(template_yolu)
    ws = wb['Veri']

    print(f"2. satirdan itibaren dolduruluyor...")
    for satir_idx, kayit in enumerate(kayitlar, start=2):
        for col_idx, anahtar in enumerate(KOLON_SIRASI, start=1):
            deger = deger_donustur(kayit.get(anahtar))
            ws.cell(row=satir_idx, column=col_idx, value=deger if deger != '' else None)

    son_satir = 1 + len(kayitlar)
    print(f"  Yazildi: 2-{son_satir}. satirlar arasi ({len(kayitlar)} kayit)")

    wb.save(template_yolu)
    print(f"OK: {template_yolu}")

    # Ozet
    kategoriler = {}
    for k in kayitlar:
        kat = k.get('kategori') or '?'
        kategoriler[kat] = kategoriler.get(kat, 0) + 1
    print()
    print('Kategori dagilimi:')
    for kat, sayi in sorted(kategoriler.items()):
        print(f"  {kat:20} {sayi:3} kayit")

if __name__ == '__main__':
    main()
