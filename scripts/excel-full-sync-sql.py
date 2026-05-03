"""
Pusula Istanbul - Excel'i tek dogruluk kaynagi olarak alir, DB'yi tam senkronize eder.

Kullanim: python3 excel-full-sync-sql.py <excel-yolu> <cikti-sql-yolu>

Mantik (DIFF DEGIL, FULL SYNC):
- Excel'deki her satir icin TUM kolonlar UPDATE'lenir
- Bos hucre = NULL (alan silinir)
- Dolu hucre = o deger
- DB'de var, Excel'de yok ise: rapor edilir (otomatik silme YOK, guvenlik)
- Excel'de mekan_id bos: yeni kayit, INSERT
"""
import json
import os
import re
import sys
import urllib.request
import urllib.parse
from openpyxl import load_workbook

KOLON_MAP = {
    'Mekan ID':           'mekan_id',
    'Isim':               'isim',
    'Tip':                'tip',
    'Kategori':           'kategori',
    'Aktif':              'aktif',
    'Mevsimsel':          'mevsimsel',
    'Restorasyon':        'restorasyon',
    'Restorasyon Notu':   'restorasyon_notu',
    'Acilis':             'acilis',
    'Kapanis':            'kapanis',
    'Gise Kapanis':       'gise_kapanis',
    'Yaz Acilis':         'yaz_acilis',
    'Yaz Kapanis':        'yaz_kapanis',
    'Yaz Gise Kapanis':   'yaz_gise_kapanis',
    'Kis Acilis':         'kis_acilis',
    'Kis Kapanis':        'kis_kapanis',
    'Kis Gise Kapanis':   'kis_gise_kapanis',
    'Kapali Gun':         'kapali_gun',
    'Haftasonu Acilis':   'haftasonu_acilis',
    'Haftasonu Kapanis':  'haftasonu_kapanis',
    'Cuma Kapali Bas':    'cuma_kapali_bas',
    'Cuma Kapali Bit':    'cuma_kapali_bit',
    'Fiyat Yerli':        'fiyat_yerli',
    'Fiyat Yabanci':      'fiyat_yabanci',
    'Fiyat Indirimli':    'fiyat_indirimli',
    'MuzeKart':           'muzekart',
    'Ozel Not':           'ozel_not',
    'Resmi Site':         'site',
    'Kaynak':             'kaynak',
}

BOOL_ALANLAR = {'aktif', 'mevsimsel', 'restorasyon'}
INT_ALANLAR = {'kapali_gun'}

# UPDATE SET'inde gecmesi gereken tum DB alanlari (mekan_id hariç — o WHERE'de)
SYNC_ALANLAR = [
    'isim', 'tip', 'kategori', 'aktif', 'mevsimsel', 'restorasyon', 'restorasyon_notu',
    'acilis', 'kapanis', 'gise_kapanis',
    'yaz_acilis', 'yaz_kapanis', 'yaz_gise_kapanis',
    'kis_acilis', 'kis_kapanis', 'kis_gise_kapanis',
    'kapali_gun', 'haftasonu_acilis', 'haftasonu_kapanis',
    'cuma_kapali_bas', 'cuma_kapali_bit',
    'fiyat_yerli', 'fiyat_yabanci', 'fiyat_indirimli', 'muzekart',
    'ozel_not', 'site', 'kaynak',
]

def env_oku(yol):
    d = {}
    if not os.path.exists(yol): return d
    with open(yol, 'r', encoding='utf-8') as f:
        for satir in f:
            satir = satir.strip()
            if not satir or satir.startswith('#') or '=' not in satir: continue
            k, v = satir.split('=', 1)
            d[k.strip()] = v.strip().strip('"').strip("'")
    return d

def db_kayitlari_cek(url, key):
    rest_url = f"{url}/rest/v1/mekan_saatleri?" + urllib.parse.urlencode({'select': 'mekan_id,isim,aktif'})
    req = urllib.request.Request(rest_url, headers={
        'apikey': key, 'Authorization': f'Bearer {key}', 'Accept': 'application/json',
    })
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode('utf-8'))

def hucre_degeri_format_ile(cell):
    """Excel currency format'ini koru: '15 €' yazilmissa value=15+format'tan birlestir."""
    val = cell.value
    if val is None or not isinstance(val, (int, float)):
        return val
    # Sayisal deger — format'ta para birimi var mi kontrol et
    fmt = (cell.number_format or '').lower()
    sembol = None
    if '€' in fmt or 'eur' in fmt: sembol = '€'
    elif '₺' in fmt or '"tl"' in fmt or '"₺"' in fmt: sembol = '₺'
    elif '$' in fmt: sembol = '$'
    elif '£' in fmt or 'gbp' in fmt: sembol = '£'
    if sembol is None:
        return val
    # Sayiyi format ile birlestir
    if isinstance(val, float) and val.is_integer():
        sayi_str = str(int(val))
    else:
        sayi_str = str(val)
    return f"{sayi_str} {sembol}"

def excel_oku(yol):
    wb = load_workbook(yol, data_only=True)
    ws = wb['Veri']
    basliklar = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    db_anahtarlari = [KOLON_MAP.get(b) for b in basliklar]

    # Fiyat alanlari currency format'i olabilir
    FIYAT_ALANLARI = {'fiyat_yerli', 'fiyat_yabanci', 'fiyat_indirimli'}

    kayitlar = []
    for satir_idx in range(2, ws.max_row + 1):
        kayit = {}
        bos_satir = True
        for col_idx, db_anahtar in enumerate(db_anahtarlari, start=1):
            if db_anahtar is None: continue
            cell = ws.cell(satir_idx, col_idx)
            if db_anahtar in FIYAT_ALANLARI:
                val = hucre_degeri_format_ile(cell)
            else:
                val = cell.value
            kayit[db_anahtar] = val
            if val is not None and (not isinstance(val, str) or val.strip() != ''):
                bos_satir = False
        if not bos_satir:
            kayit['_satir_no'] = satir_idx
            kayitlar.append(kayit)
    return kayitlar

def excel_degeri_normalize(deger, alan):
    if deger is None: return None
    if isinstance(deger, str):
        s = deger.strip()
        if s == '': return None
        if alan in BOOL_ALANLAR:
            if s.upper() in ('TRUE', '1', 'YES', 'EVET'): return True
            if s.upper() in ('FALSE', '0', 'NO', 'HAYIR'): return False
        if alan in INT_ALANLAR:
            try: return int(s)
            except: return s
        return s
    if isinstance(deger, bool): return deger
    if isinstance(deger, int) and alan in INT_ALANLAR: return deger
    if isinstance(deger, float):
        if alan in INT_ALANLAR: return int(deger)
        if deger.is_integer(): return str(int(deger))
        return str(deger)
    return deger

def sql_value(deger, alan):
    if deger is None: return 'NULL'
    if alan in BOOL_ALANLAR:
        if isinstance(deger, bool): return 'TRUE' if deger else 'FALSE'
    if alan in INT_ALANLAR:
        return str(int(deger))
    s = str(deger).replace("'", "''")
    return f"'{s}'"

def kayit_normallestir(kayit):
    """NOT NULL constraint'lar icin akilli varsayilanlar:
    - kapanis bos VE gise_kapanis dolu ise: kapanis = gise_kapanis
    """
    kapanis_raw = kayit.get('kapanis')
    kapanis_norm = excel_degeri_normalize(kapanis_raw, 'kapanis')
    if kapanis_norm is None:
        gise_norm = excel_degeri_normalize(kayit.get('gise_kapanis'), 'gise_kapanis')
        if gise_norm is not None:
            kayit['kapanis'] = gise_norm
    return kayit

def update_sql_full(mekan_id, kayit):
    """Full sync UPDATE — TUM SYNC_ALANLAR'i SET et, bosturanlari NULL yap."""
    kayit = kayit_normallestir(dict(kayit))
    set_parcalari = []
    for alan in SYNC_ALANLAR:
        deger = excel_degeri_normalize(kayit.get(alan), alan)
        set_parcalari.append(f"  {alan} = {sql_value(deger, alan)}")
    set_parcalari.append(f"  guncelleme_tarihi = NOW()")
    safe_id = mekan_id.replace("'", "''")
    return "UPDATE mekan_saatleri SET\n" + ",\n".join(set_parcalari) + f"\nWHERE mekan_id = '{safe_id}';"

def insert_sql(kayit):
    """Yeni kayit INSERT — Excel'deki tum alanlari yaz."""
    if not kayit.get('mekan_id'):
        if not kayit.get('isim'):
            return None, "isim bos, mekan_id uretilemez"
        slug = kayit['isim'].lower()
        slug = re.sub(r'[^a-z0-9ğüşıöç]+', '_', slug)
        kayit['mekan_id'] = slug.strip('_')
    kayit.setdefault('aktif_mevsim', 'kis')

    alanlar = ['mekan_id'] + SYNC_ALANLAR + ['aktif_mevsim']
    degerler = []
    for alan in alanlar:
        deger = excel_degeri_normalize(kayit.get(alan), alan) if alan != 'aktif_mevsim' else kayit.get(alan)
        degerler.append(sql_value(deger, alan))
    return f"INSERT INTO mekan_saatleri ({', '.join(alanlar)})\nVALUES ({', '.join(degerler)});", None

def main():
    if len(sys.argv) < 3:
        print('Kullanim: python3 excel-full-sync-sql.py <excel-yolu> <cikti-sql-yolu>'); sys.exit(1)

    excel_yolu = sys.argv[1]
    cikti_yolu = sys.argv[2]
    proje_kok = '/sessions/magical-upbeat-mccarthy/mnt/istanbul-rehber'
    env = env_oku(os.path.join(proje_kok, '.env'))
    url = env.get('SUPABASE_URL') or 'https://rzlfghjpsximthlolfxo.supabase.co'
    key = env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not key:
        print("HATA: SUPABASE_SERVICE_ROLE_KEY bulunamadi"); sys.exit(1)

    print(f"Excel okunuyor: {excel_yolu}")
    excel_kayitlari = excel_oku(excel_yolu)
    print(f"  {len(excel_kayitlari)} satir okundu")

    print(f"DB'den kayit listesi cekiliyor (sadece kontrol icin)...")
    db_kayitlari = db_kayitlari_cek(url, key)
    db_idler = {k['mekan_id'] for k in db_kayitlari}
    print(f"  {len(db_kayitlari)} kayit DB'de var")

    update_sqlleri = []
    insert_sqlleri = []
    excel_idleri = set()

    for excel_kayit in excel_kayitlari:
        mekan_id = excel_kayit.get('mekan_id')
        if mekan_id and mekan_id in db_idler:
            update_sqlleri.append(update_sql_full(mekan_id, excel_kayit))
            excel_idleri.add(mekan_id)
        else:
            kayit_kopya = {k: v for k, v in excel_kayit.items() if not k.startswith('_')}
            sql, hata = insert_sql(kayit_kopya)
            if sql:
                insert_sqlleri.append(sql)
                if mekan_id: excel_idleri.add(mekan_id)

    db_de_var_excel_de_yok = db_idler - excel_idleri

    with open(cikti_yolu, 'w', encoding='utf-8') as f:
        f.write("-- Pusula Istanbul - Mekan Saatleri FULL SYNC (Excel = tek dogruluk kaynagi)\n")
        f.write(f"-- Kaynak: {os.path.basename(excel_yolu)}\n")
        f.write(f"-- Mod: FULL SYNC (bos hucre = NULL, dolu hucre = o deger)\n")
        f.write(f"-- {len(update_sqlleri)} UPDATE + {len(insert_sqlleri)} INSERT\n\n")
        f.write("BEGIN;\n\n")
        if update_sqlleri:
            f.write("-- ═══════════════ UPDATE'ler (full sync) ═══════════════\n\n")
            for sql in update_sqlleri:
                f.write(sql + "\n\n")
        if insert_sqlleri:
            f.write("-- ═══════════════ INSERT'ler ═══════════════\n\n")
            for sql in insert_sqlleri:
                f.write(sql + "\n\n")
        f.write("COMMIT;\n")
        if db_de_var_excel_de_yok:
            f.write(f"\n-- ═══════════════ UYARI ═══════════════\n")
            f.write(f"-- DB'de var ama Excel'de yok ({len(db_de_var_excel_de_yok)} kayit):\n")
            for mid in sorted(db_de_var_excel_de_yok):
                f.write(f"--   {mid}\n")
            f.write(f"-- Bunlar otomatik SILINMEDI. Eger silinmesi gerekiyorsa manuel DELETE statement yazilmali.\n")

    print()
    print("═" * 70)
    print(f"OZET (FULL SYNC modu)")
    print("═" * 70)
    print(f"  UPDATE: {len(update_sqlleri)} kayit (her biri TUM alanlari overwrite eder)")
    print(f"  INSERT: {len(insert_sqlleri)} kayit")
    if db_de_var_excel_de_yok:
        print(f"  DB'de var, Excel'de yok: {len(db_de_var_excel_de_yok)} kayit (silinmedi, manuel karar)")
        for mid in sorted(db_de_var_excel_de_yok):
            print(f"    - {mid}")
    print()
    print(f"SQL dosyasi: {cikti_yolu}")

if __name__ == '__main__':
    main()
