#!/usr/bin/env python3
"""
Pusula Istanbul — Abone / Freemium Listesi Excel Ureticisi
----------------------------------------------------------
Supabase'den tum musteri kayitlarini ceker ve 4 sekmeli formatli bir .xlsx
uretir. Hem elle hem gunluk scheduled task ile calistirilabilir; her calismada
dosyayi bastan uretir (idempotent).

ODEME KRITERI: revenuecat_id DOLU = en az bir kez gercek odeme yapmis.
Promosyon / hediye ile premium olmuslar (revenuecat_id BOS) FREEMIUM sayilir.
Personel (admin + moderator) tum sekmelerin disinda.

Sekmeler:
  1. Yillik Aboneler        — odeyen + durumu aktif + plan yillik + bitis gelecekte
  2. Aylik Aboneler         — odeyen + durumu aktif + plan aylik  + bitis gelecekte
  3. Abonelikleri Bitmisler — odeyen ama su an aktif premium degil (bitmis/iptal)
  4. Freemium               — revenuecat_id BOS (hic odeme yok / promosyon)

Calistirma:
  python3 scripts/aboneler-excel.py

Servis anahtari .env dosyasindaki SUPABASE_SERVICE_ROLE_KEY'den okunur.
"""
import sys
import json
import urllib.request
import urllib.parse
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SUPABASE_URL = "https://rzlfghjpsximthlolfxo.supabase.co"
PROJE_KOK = Path(__file__).resolve().parent.parent
CIKTI = PROJE_KOK / "aboneler-premium.xlsx"

MAVI = "0077B6"
KOYU_MAVI = "005A8D"
ACIK_MAVI = "CDEFFB"
ALTIN = "C77A15"
ZEBRA = "F2F8FC"


def service_key() -> str:
    env = PROJE_KOK / ".env"
    if not env.exists():
        sys.exit(".env bulunamadi: " + str(env))
    for satir in env.read_text(encoding="utf-8").splitlines():
        satir = satir.strip()
        if satir.startswith("SUPABASE_SERVICE_ROLE_KEY"):
            return satir.split("=", 1)[1].strip().strip('"').strip("'")
    sys.exit(".env icinde SUPABASE_SERVICE_ROLE_KEY yok")


def musterileri_cek() -> list:
    params = (
        "select=isim,soyisim,email,abonelik_plani,abonelik_durumu,abonelik_bitis,revenuecat_id,rol"
        "&rol=not.in.(admin,moderator)"
        "&order=abonelik_bitis.desc"
    )
    url = f"{SUPABASE_URL}/rest/v1/profiles?{params}"
    key = service_key()
    req = urllib.request.Request(url, headers={
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
        "Prefer": "count=none",
        "Range": "0-9999",
    })
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode("utf-8"))


def tr_tarih(iso):
    if not iso:
        return None
    s = iso.replace(" ", "T")
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        dt = datetime.fromisoformat(s.split("+")[0])
    return dt.replace(tzinfo=None)  # Excel timezone'lu datetime kabul etmez


def odemis(k) -> bool:
    rc = k.get("revenuecat_id")
    return rc is not None and str(rc).strip() != ""


def aktif_premium(k, now) -> bool:
    dt = tr_tarih(k.get("abonelik_bitis"))
    return k.get("abonelik_durumu") == "aktif" and dt is not None and dt > now


INCE = Side(style="thin", color="D0D7DE")
KENAR = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)


def sekme_yaz(ws, baslik, kayitlar, kolonlar, uretim):
    """kolonlar: [(baslik, genislik, deger_fn, hizalama)]"""
    n = len(kolonlar)
    son_harf = chr(ord('A') + n - 1)

    ws.merge_cells(f"A1:{son_harf}1")
    b = ws["A1"]
    b.value = baslik
    b.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    b.fill = PatternFill("solid", fgColor=MAVI)
    b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{son_harf}2")
    a = ws["A2"]
    a.value = f"Guncelleme: {uretim:%d.%m.%Y %H:%M} (TR)   ·   Kayit: {len(kayitlar)}"
    a.font = Font(name="Arial", size=10, color="444444")
    a.fill = PatternFill("solid", fgColor=ACIK_MAVI)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    hr = 4
    for i, (kbaslik, _, _, _) in enumerate(kolonlar, start=1):
        c = ws.cell(row=hr, column=i, value=kbaslik)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=KOYU_MAVI)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = KENAR
    ws.row_dimensions[hr].height = 22

    for idx, k in enumerate(kayitlar, start=1):
        r = hr + idx
        for ci, (_, _, fn, hiz) in enumerate(kolonlar, start=1):
            val = fn(k, idx)
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal=hiz, vertical="center")
            cell.border = KENAR
            if isinstance(val, datetime):
                cell.number_format = "dd.mm.yyyy"
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)

    for i, (_, w, _, _) in enumerate(kolonlar):
        ws.column_dimensions[chr(ord('A') + i)].width = w

    ws.freeze_panes = f"A{hr + 1}"
    if kayitlar:
        ws.auto_filter.ref = f"A{hr}:{son_harf}{hr + len(kayitlar)}"


def uret():
    musteriler = musterileri_cek()
    now = datetime.now(timezone.utc).replace(tzinfo=None)  # tr_tarih de naive donuyor
    uretim = now + timedelta(hours=3)  # TR (DST yok)

    yillik, aylik, bitmis, freemium = [], [], [], []
    for k in musteriler:
        if not odemis(k):
            freemium.append(k)
        elif aktif_premium(k, now):
            (yillik if k.get("abonelik_plani") == "yillik" else aylik).append(k)
        else:
            bitmis.append(k)

    # siralamalar
    yillik.sort(key=lambda k: k.get("abonelik_bitis") or "")
    aylik.sort(key=lambda k: k.get("abonelik_bitis") or "")
    bitmis.sort(key=lambda k: k.get("abonelik_bitis") or "", reverse=True)
    freemium.sort(key=lambda k: ((k.get("isim") or "").lower(), (k.get("soyisim") or "").lower()))

    ad = lambda k, i: (k.get("isim") or "").strip()
    soyad = lambda k, i: (k.get("soyisim") or "").strip()
    mail = lambda k, i: k.get("email")
    sira = lambda k, i: i
    plan_tr = {"aylik": "Aylik", "yillik": "Yillik"}
    plan = lambda k, i: plan_tr.get(k.get("abonelik_plani"), k.get("abonelik_plani") or "-")
    bitis = lambda k, i: (tr_tarih(k.get("abonelik_bitis")) or "-")
    durum = lambda k, i: k.get("abonelik_durumu") or "-"

    abone_kol = [
        ("Sira", 6, sira, "center"),
        ("Ad", 16, ad, "left"),
        ("Soyad", 18, soyad, "left"),
        ("E-posta", 34, mail, "left"),
        ("Plan", 9, plan, "center"),
        ("Bitis Tarihi", 14, bitis, "center"),
    ]
    bitmis_kol = [
        ("Sira", 6, sira, "center"),
        ("Ad", 16, ad, "left"),
        ("Soyad", 18, soyad, "left"),
        ("E-posta", 34, mail, "left"),
        ("Son Plan", 9, plan, "center"),
        ("Bitis Tarihi", 14, bitis, "center"),
        ("Durum", 14, durum, "center"),
    ]
    freemium_kol = [
        ("Sira", 6, sira, "center"),
        ("Ad", 16, ad, "left"),
        ("Soyad", 18, soyad, "left"),
        ("E-posta", 34, mail, "left"),
        ("Durum", 16, durum, "center"),
    ]

    wb = Workbook()
    sekmeler = [
        ("Yillik Aboneler", f"Yillik Aboneler (odeyen) — {len(yillik)} kisi", yillik, abone_kol),
        ("Aylik Aboneler", f"Aylik Aboneler (odeyen) — {len(aylik)} kisi", aylik, abone_kol),
        ("Abonelikleri Bitmisler", f"Abonelikleri Bitmisler (odemis, su an pasif) — {len(bitmis)} kisi", bitmis, bitmis_kol),
        ("Freemium", f"Freemium (odeme yok / promosyon) — {len(freemium)} kisi", freemium, freemium_kol),
    ]
    for i, (sekme_ad, baslik, kayitlar, kol) in enumerate(sekmeler):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sekme_ad
        sekme_yaz(ws, baslik, kayitlar, kol, uretim)

    wb.save(CIKTI)
    print(json.dumps({
        "dosya": str(CIKTI),
        "yillik": len(yillik), "aylik": len(aylik),
        "bitmis": len(bitmis), "freemium": len(freemium),
        "guncelleme": uretim.strftime("%d.%m.%Y %H:%M")
    }, ensure_ascii=False))


if __name__ == "__main__":
    uret()
